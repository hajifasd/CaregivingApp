"""
护工资源管理系统 - 报告导出器
====================================

支持Excel和PDF格式的报告导出功能
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, Any, Optional
import io

# 安全导入可选依赖
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

logger = logging.getLogger(__name__)

class ReportExporter:
    """报告导出器"""
    
    def __init__(self):
        self.has_pandas = HAS_PANDAS
        self.has_openpyxl = HAS_OPENPYXL
        self.has_reportlab = HAS_REPORTLAB
        
        # 检查依赖
        if not self.has_pandas:
            logger.warning("⚠️ pandas未安装，Excel导出功能将受限")
        if not self.has_openpyxl:
            logger.warning("⚠️ openpyxl未安装，Excel导出功能将不可用")
        if not self.has_reportlab:
            logger.warning("⚠️ reportlab未安装，PDF导出功能将不可用")
    
    def export_to_excel(self, analysis_data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """导出分析结果为Excel格式"""
        try:
            if not self.has_openpyxl:
                raise ImportError("openpyxl未安装，无法导出Excel文件")
            
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"护工数据分析报告_{timestamp}.xlsx"
            
            # 创建工作簿
            wb = Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 创建样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. 概览信息工作表
            self._create_overview_sheet(wb, analysis_data, header_font, header_fill, border)
            
            # 2. 薪资分析工作表
            if 'salary_analysis' in analysis_data:
                self._create_salary_analysis_sheet(wb, analysis_data['salary_analysis'], 
                                                 header_font, header_fill, border)
            
            # 3. 技能分析工作表
            if 'skill_analysis' in analysis_data:
                self._create_skill_analysis_sheet(wb, analysis_data['skill_analysis'], 
                                                header_font, header_fill, border)
            
            # 4. 护工分布分析工作表
            if 'caregiver_distribution' in analysis_data:
                self._create_distribution_sheet(wb, analysis_data['caregiver_distribution'], 
                                              header_font, header_fill, border)
            
            # 3. 成功率预测工作表
            if 'success_prediction' in analysis_data:
                self._create_prediction_sheet(wb, analysis_data['success_prediction'], 
                                            header_font, header_fill, border)
            
            # 4. 聚类分析工作表
            if 'cluster_analysis' in analysis_data:
                self._create_cluster_sheet(wb, analysis_data['cluster_analysis'], 
                                         header_font, header_fill, border)
            
            # 5. 市场趋势工作表
            if 'market_trends' in analysis_data:
                self._create_trends_sheet(wb, analysis_data['market_trends'], 
                                        header_font, header_fill, border)
            
            # 6. 需求预测工作表
            if 'demand_forecast' in analysis_data:
                self._create_forecast_sheet(wb, analysis_data['demand_forecast'], 
                                          header_font, header_fill, border)
            
            # 保存文件
            # 从项目根目录开始计算路径
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
            exports_dir = os.path.join(project_root, 'exports')
            os.makedirs(exports_dir, exist_ok=True)
            filepath = os.path.join(exports_dir, filename)
            wb.save(filepath)
            
            logger.info(f"✅ Excel报告导出成功: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"❌ Excel报告导出失败: {str(e)}")
            raise
    
    def export_to_pdf(self, analysis_data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """导出分析结果为PDF格式"""
        try:
            if not self.has_reportlab:
                raise ImportError("reportlab未安装，无法导出PDF文件")
            
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"护工数据分析报告_{timestamp}.pdf"
            
            # 创建PDF文档
            # 从项目根目录开始计算路径
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
            exports_dir = os.path.join(project_root, 'exports')
            os.makedirs(exports_dir, exist_ok=True)
            filepath = os.path.join(exports_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # 获取样式
            styles = getSampleStyleSheet()
            
            # 创建自定义样式
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            
            # 1. 标题页
            story.append(Paragraph("护工资源管理系统", title_style))
            story.append(Paragraph("大数据分析报告", title_style))
            story.append(Spacer(1, 20))
            
            # 添加报告信息
            report_info = [
                ["报告生成时间", analysis_data.get('analysis_timestamp', '未知')],
                ["数据版本", str(analysis_data.get('data_version', '未知'))],
                ["触发源", analysis_data.get('trigger_source', '未知')],
                ["数据源", ', '.join(analysis_data.get('data_sources', ['未知']))]
            ]
            
            info_table = Table(report_info, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(info_table)
            story.append(PageBreak())
            
            # 2. 薪资分析
            if 'salary_analysis' in analysis_data:
                self._add_salary_analysis_to_pdf(story, analysis_data['salary_analysis'], 
                                               heading_style, styles)
            
            # 3. 技能分析
            if 'skill_analysis' in analysis_data:
                self._add_skill_analysis_to_pdf(story, analysis_data['skill_analysis'], 
                                              heading_style, styles)
            
            # 4. 护工分布分析
            if 'caregiver_distribution' in analysis_data:
                self._add_distribution_to_pdf(story, analysis_data['caregiver_distribution'], 
                                            heading_style, styles)
            
            # 3. 成功率预测
            if 'success_prediction' in analysis_data:
                self._add_prediction_to_pdf(story, analysis_data['success_prediction'], 
                                          heading_style, styles)
            
            # 4. 聚类分析
            if 'cluster_analysis' in analysis_data:
                self._add_cluster_to_pdf(story, analysis_data['cluster_analysis'], 
                                       heading_style, styles)
            
            # 5. 市场趋势
            if 'market_trends' in analysis_data:
                self._add_trends_to_pdf(story, analysis_data['market_trends'], 
                                      heading_style, styles)
            
            # 6. 需求预测
            if 'demand_forecast' in analysis_data:
                self._add_forecast_to_pdf(story, analysis_data['demand_forecast'], 
                                        heading_style, styles)
            
            # 生成PDF
            doc.build(story)
            
            logger.info(f"✅ PDF报告导出成功: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"❌ PDF报告导出失败: {str(e)}")
            raise
    
    def _create_overview_sheet(self, wb, analysis_data, header_font, header_fill, border):
        """创建概览信息工作表"""
        ws = wb.create_sheet("概览信息")
        
        # 标题
        ws['A1'] = "护工资源管理系统 - 大数据分析报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # 报告信息
        row = 3
        info_data = [
            ["报告生成时间", analysis_data.get('analysis_timestamp', '未知')],
            ["数据版本", str(analysis_data.get('data_version', '未知'))],
            ["触发源", analysis_data.get('trigger_source', '未知')],
            ["数据源", ', '.join(analysis_data.get('data_sources', ['未知']))]
        ]
        
        for info in info_data:
            ws[f'A{row}'] = info[0]
            ws[f'B{row}'] = info[1]
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # 数据统计
        if 'data_counts' in analysis_data:
            row += 2
            ws[f'A{row}'] = "数据统计"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            for source, count in analysis_data['data_counts'].items():
                ws[f'A{row}'] = source
                ws[f'B{row}'] = count
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_salary_analysis_sheet(self, wb, salary_data, header_font, header_fill, border):
        """创建薪资分析工作表"""
        ws = wb.create_sheet("薪资分析")
        
        # 标题
        ws['A1'] = "薪资分析报告"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # 总体统计
        if 'total_avg' in salary_data:
            ws[f'A{row}'] = "平均薪资"
            ws[f'B{row}'] = f"{salary_data['total_avg']:.2f} 元/月"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        if 'total_records' in salary_data:
            ws[f'A{row}'] = "数据记录数"
            ws[f'B{row}'] = salary_data['total_records']
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 2
        
        # 城市薪资分布
        if 'city_avg' in salary_data:
            ws[f'A{row}'] = "城市薪资分布"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = "平均薪资 (元/月)"
            ws[f'B{row}'].font = header_font
            ws[f'B{row}'].fill = header_fill
            row += 1
            
            for city, salary in salary_data['city_avg'].items():
                ws[f'A{row}'] = city
                ws[f'B{row}'] = f"{salary:.2f}"
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def _create_skill_analysis_sheet(self, wb, skill_data, header_font, header_fill, border):
        """创建技能分析工作表"""
        ws = wb.create_sheet("技能分析")
        
        # 标题
        ws['A1'] = "技能需求分析报告"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # 总体统计
        if 'total_records' in skill_data:
            ws[f'A{row}'] = "数据记录数"
            ws[f'B{row}'] = skill_data['total_records']
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 2
        
        # 技能需求统计
        if 'skill_counts' in skill_data:
            ws[f'A{row}'] = "技能需求统计"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = "需求次数"
            ws[f'B{row}'].font = header_font
            ws[f'B{row}'].fill = header_fill
            row += 1
            
            # 按需求次数排序
            sorted_skills = sorted(skill_data['skill_counts'].items(), 
                                 key=lambda x: x[1], reverse=True)
            
            for skill, count in sorted_skills:
                ws[f'A{row}'] = skill
                ws[f'B{row}'] = count
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_distribution_sheet(self, wb, distribution_data, header_font, header_fill, border):
        """创建护工分布分析工作表"""
        ws = wb.create_sheet("护工分布分析")
        
        # 标题
        ws['A1'] = "护工分布分析"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # 地域分布
        if 'location_distribution' in distribution_data:
            ws[f'A{row}'] = "地域分布"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = "护工数量"
            ws[f'B{row}'].font = header_font
            ws[f'B{row}'].fill = header_fill
            row += 1
            
            for location, count in distribution_data['location_distribution'].items():
                ws[f'A{row}'] = location
                ws[f'B{row}'] = count
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        row += 2
        
        # 薪资分布
        if 'salary_distribution' in distribution_data:
            ws[f'A{row}'] = "薪资分布"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = "护工数量"
            ws[f'B{row}'].font = header_font
            ws[f'B{row}'].fill = header_fill
            row += 1
            
            for salary_range, count in distribution_data['salary_distribution'].items():
                ws[f'A{row}'] = salary_range
                ws[f'B{row}'] = count
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_prediction_sheet(self, wb, prediction_data, header_font, header_fill, border):
        """创建成功率预测工作表"""
        ws = wb.create_sheet("成功率预测")
        
        # 标题
        ws['A1'] = "护工成功率预测分析"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # 模型准确率
        ws[f'A{row}'] = "模型准确率"
        ws[f'B{row}'] = prediction_data.get('model_accuracy', 0)
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 2
        
        # 特征重要性
        if 'feature_importance' in prediction_data:
            ws[f'A{row}'] = "特征重要性"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = "重要性得分"
            ws[f'B{row}'].font = header_font
            ws[f'B{row}'].fill = header_fill
            row += 1
            
            for feature, importance in prediction_data['feature_importance'].items():
                ws[f'A{row}'] = feature
                ws[f'B{row}'] = importance
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_cluster_sheet(self, wb, cluster_data, header_font, header_fill, border):
        """创建聚类分析工作表"""
        ws = wb.create_sheet("聚类分析")
        
        # 标题
        ws['A1'] = "护工聚类分析"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        row = 3
        
        # 表头
        headers = ["聚类ID", "护工数量", "平均年龄", "平均时薪", "平均评分"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        
        # 聚类数据
        if 'clusters' in cluster_data:
            for cluster in cluster_data['clusters']:
                ws[f'A{row}'] = cluster.get('cluster_id', 0)
                ws[f'B{row}'] = cluster.get('count', 0)
                ws[f'C{row}'] = cluster.get('avg_age', 0)
                ws[f'D{row}'] = cluster.get('avg_hourly_rate', 0)
                ws[f'E{row}'] = cluster.get('avg_rating', 0)
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def _create_trends_sheet(self, wb, trends_data, header_font, header_fill, border):
        """创建市场趋势工作表"""
        ws = wb.create_sheet("市场趋势")
        
        # 标题
        ws['A1'] = "市场趋势分析"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        row = 3
        
        # 表头
        headers = ["数据源", "平均薪资", "职位数量", "最低薪资", "最高薪资"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        
        # 趋势数据
        if 'trends' in trends_data:
            for trend in trends_data['trends']:
                ws[f'A{row}'] = trend.get('data_source', '')
                ws[f'B{row}'] = trend.get('avg_salary', 0)
                ws[f'C{row}'] = trend.get('job_count', 0)
                ws[f'D{row}'] = trend.get('min_salary', 0)
                ws[f'E{row}'] = trend.get('max_salary', 0)
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def _create_forecast_sheet(self, wb, forecast_data, header_font, header_fill, border):
        """创建需求预测工作表"""
        ws = wb.create_sheet("需求预测")
        
        # 标题
        ws['A1'] = "需求预测分析"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        row = 3
        
        # 模型准确率
        ws[f'A{row}'] = "模型准确率"
        ws[f'B{row}'] = forecast_data.get('model_accuracy', 0)
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 3
        
        # 表头
        headers = ["日期", "预测需求", "置信区间下限", "置信区间上限", "趋势"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        
        # 预测数据
        if 'forecast' in forecast_data:
            for forecast in forecast_data['forecast']:
                ws[f'A{row}'] = forecast.get('date', '')
                ws[f'B{row}'] = forecast.get('predicted_demand', 0)
                ws[f'C{row}'] = forecast.get('confidence_interval', {}).get('lower', 0)
                ws[f'D{row}'] = forecast.get('confidence_interval', {}).get('upper', 0)
                ws[f'E{row}'] = forecast.get('trend', '')
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def _add_salary_analysis_to_pdf(self, story, salary_data, heading_style, styles):
        """添加薪资分析到PDF"""
        story.append(Paragraph("薪资分析", heading_style))
        
        # 总体统计
        if 'total_avg' in salary_data:
            story.append(Paragraph(f"平均薪资: {salary_data['total_avg']:.2f} 元/月", styles['Normal']))
        
        if 'total_records' in salary_data:
            story.append(Paragraph(f"数据记录数: {salary_data['total_records']}", styles['Normal']))
        
        story.append(Spacer(1, 12))
        
        # 城市薪资分布
        if 'city_avg' in salary_data:
            story.append(Paragraph("城市薪资分布", styles['Heading3']))
            
            salary_table_data = [["城市", "平均薪资 (元/月)"]]
            for city, salary in salary_data['city_avg'].items():
                salary_table_data.append([city, f"{salary:.2f}"])
            
            salary_table = Table(salary_table_data)
            salary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(salary_table)
            story.append(Spacer(1, 12))
    
    def _add_skill_analysis_to_pdf(self, story, skill_data, heading_style, styles):
        """添加技能分析到PDF"""
        story.append(Paragraph("技能需求分析", heading_style))
        
        # 总体统计
        if 'total_records' in skill_data:
            story.append(Paragraph(f"数据记录数: {skill_data['total_records']}", styles['Normal']))
        
        story.append(Spacer(1, 12))
        
        # 技能需求统计
        if 'skill_counts' in skill_data:
            story.append(Paragraph("技能需求统计", styles['Heading3']))
            
            skill_table_data = [["技能", "需求次数"]]
            # 按需求次数排序
            sorted_skills = sorted(skill_data['skill_counts'].items(), 
                                 key=lambda x: x[1], reverse=True)
            
            for skill, count in sorted_skills:
                skill_table_data.append([skill, str(count)])
            
            skill_table = Table(skill_table_data)
            skill_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(skill_table)
            story.append(Spacer(1, 12))
    
    def _add_distribution_to_pdf(self, story, distribution_data, heading_style, styles):
        """添加护工分布分析到PDF"""
        story.append(Paragraph("护工分布分析", heading_style))
        
        # 地域分布
        if 'location_distribution' in distribution_data:
            story.append(Paragraph("地域分布", styles['Heading3']))
            
            location_data = [["城市", "护工数量"]]
            for location, count in distribution_data['location_distribution'].items():
                location_data.append([location, str(count)])
            
            location_table = Table(location_data)
            location_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(location_table)
            story.append(Spacer(1, 12))
        
        # 薪资分布
        if 'salary_distribution' in distribution_data:
            story.append(Paragraph("薪资分布", styles['Heading3']))
            
            salary_data = [["薪资范围", "护工数量"]]
            for salary_range, count in distribution_data['salary_distribution'].items():
                salary_data.append([salary_range, str(count)])
            
            salary_table = Table(salary_data)
            salary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(salary_table)
            story.append(Spacer(1, 12))
    
    def _add_prediction_to_pdf(self, story, prediction_data, heading_style, styles):
        """添加成功率预测到PDF"""
        story.append(Paragraph("成功率预测分析", heading_style))
        
        # 模型准确率
        accuracy = prediction_data.get('model_accuracy', 0)
        story.append(Paragraph(f"模型准确率: {accuracy:.3f}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # 特征重要性
        if 'feature_importance' in prediction_data:
            story.append(Paragraph("特征重要性", styles['Heading3']))
            
            feature_data = [["特征", "重要性得分"]]
            for feature, importance in prediction_data['feature_importance'].items():
                feature_data.append([feature, f"{importance:.3f}"])
            
            feature_table = Table(feature_data)
            feature_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(feature_table)
            story.append(Spacer(1, 12))
    
    def _add_cluster_to_pdf(self, story, cluster_data, heading_style, styles):
        """添加聚类分析到PDF"""
        story.append(Paragraph("聚类分析", heading_style))
        
        if 'clusters' in cluster_data:
            cluster_table_data = [["聚类ID", "护工数量", "平均年龄", "平均时薪", "平均评分"]]
            
            for cluster in cluster_data['clusters']:
                cluster_table_data.append([
                    str(cluster.get('cluster_id', 0)),
                    str(cluster.get('count', 0)),
                    str(cluster.get('avg_age', 0)),
                    str(cluster.get('avg_hourly_rate', 0)),
                    str(cluster.get('avg_rating', 0))
                ])
            
            cluster_table = Table(cluster_table_data)
            cluster_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(cluster_table)
            story.append(Spacer(1, 12))
    
    def _add_trends_to_pdf(self, story, trends_data, heading_style, styles):
        """添加市场趋势到PDF"""
        story.append(Paragraph("市场趋势分析", heading_style))
        
        if 'trends' in trends_data:
            trend_table_data = [["数据源", "平均薪资", "职位数量", "最低薪资", "最高薪资"]]
            
            for trend in trends_data['trends']:
                trend_table_data.append([
                    str(trend.get('data_source', '')),
                    str(trend.get('avg_salary', 0)),
                    str(trend.get('job_count', 0)),
                    str(trend.get('min_salary', 0)),
                    str(trend.get('max_salary', 0))
                ])
            
            trend_table = Table(trend_table_data)
            trend_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(trend_table)
            story.append(Spacer(1, 12))
    
    def _add_forecast_to_pdf(self, story, forecast_data, heading_style, styles):
        """添加需求预测到PDF"""
        story.append(Paragraph("需求预测分析", heading_style))
        
        # 模型准确率
        accuracy = forecast_data.get('model_accuracy', 0)
        story.append(Paragraph(f"模型准确率: {accuracy:.3f}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        if 'forecast' in forecast_data:
            forecast_table_data = [["日期", "预测需求", "置信区间下限", "置信区间上限", "趋势"]]
            
            for forecast in forecast_data['forecast']:
                forecast_table_data.append([
                    str(forecast.get('date', '')),
                    str(forecast.get('predicted_demand', 0)),
                    str(forecast.get('confidence_interval', {}).get('lower', 0)),
                    str(forecast.get('confidence_interval', {}).get('upper', 0)),
                    str(forecast.get('trend', ''))
                ])
            
            forecast_table = Table(forecast_table_data)
            forecast_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(forecast_table)
            story.append(Spacer(1, 12))

def main():
    """主函数 - 示例用法"""
    exporter = ReportExporter()
    
    # 模拟分析数据
    mock_data = {
        'analysis_timestamp': datetime.now().isoformat(),
        'data_version': 1,
        'trigger_source': 'test',
        'data_sources': ['test_data'],
        'data_counts': {'test_data': 100},
        'caregiver_distribution': {
            'location_distribution': {'北京': 50, '上海': 40, '广州': 30},
            'salary_distribution': {'3000-5000': 60, '5000-8000': 30, '8000+': 10},
            'total_caregivers': 100
        },
        'success_prediction': {
            'model_accuracy': 0.85,
            'feature_importance': {'rating': 0.35, 'experience': 0.25, 'salary': 0.20}
        }
    }
    
    try:
        # 导出Excel
        excel_file = exporter.export_to_excel(mock_data)
        print(f"✅ Excel报告导出成功: {excel_file}")
        
        # 导出PDF
        pdf_file = exporter.export_to_pdf(mock_data)
        print(f"✅ PDF报告导出成功: {pdf_file}")
        
    except Exception as e:
        print(f"❌ 导出失败: {str(e)}")

if __name__ == "__main__":
    main()
